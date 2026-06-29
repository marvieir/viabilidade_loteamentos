"""Export Excel (.xlsx) das dimensões já executadas (Fase Tier-1, paridade Urbia "Download Excel").

Recebe a identificação + os dicts das dimensões (o mesmo corpo do laudo PDF; o front repassa o
que o backend devolveu — nada recalculado). Monta uma planilha por bloco relevante. Offline,
sem rede, determinístico. Defensivo: dimensão/campo ausente → aba/linha omitida (não quebra)."""

from __future__ import annotations

import io
from typing import Any, Optional


def _num(v: Any) -> Any:
    """Devolve número (int/float) se for numérico; senão o valor original (string)."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return v


def _g(d: Any, *path: str, default: Any = None) -> Any:
    cur = d
    for k in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
    return cur if cur is not None else default


def gerar_excel(ident: dict, dims: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    negrito = Font(bold=True)

    def nova_aba(titulo: str):
        ws = wb.create_sheet(title=titulo[:31])
        return ws

    def kv(ws, rotulo: str, valor: Any):
        ws.append([rotulo, _num(valor)])

    def cabecalho(ws, colunas: list[str]):
        ws.append(colunas)
        for c in ws[ws.max_row]:
            c.font = negrito

    # ---- Resumo (sempre) ----
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Pré-Viabilidade de Loteamento — Resumo"])
    ws["A1"].font = negrito
    ws.append([])
    kv(ws, "Análise", ident.get("analise_id"))
    kv(ws, "Município / UF", f"{ident.get('municipio') or '—'} / {ident.get('uf') or '—'}")
    kv(ws, "Área (m²)", ident.get("area_m2"))
    kv(ws, "Área (ha)", ident.get("area_ha"))
    kv(ws, "Perímetro (m)", ident.get("perimetro_m"))
    apr = dims.get("aproveitamento") or {}
    for rot, chaves in [
        ("Aproveitável (m²)", ("aproveitavel_m2", "area_aproveitavel_m2")),
        ("Aproveitável (ha)", ("aproveitavel_ha",)),
        ("Nº de lotes (teto físico)", ("n_lotes", "lotes", "n_lotes_teto")),
    ]:
        val = next((apr.get(k) for k in chaves if apr.get(k) is not None), None)
        if val is not None:
            kv(ws, rot, val)

    # ---- Declividade (faixas finas) ----
    decl = dims.get("declividade")
    if isinstance(decl, dict) and decl.get("consultada"):
        ws = nova_aba("Declividade")
        kv(ws, "Declividade média (%)", decl.get("declividade_media_pct"))
        kv(ws, "Relevo predominante", decl.get("relevo_predominante"))
        ws.append([])
        cabecalho(ws, ["Faixa", "Área (m²)", "% da gleba"])
        for f in decl.get("faixas_finas") or []:
            if (f.get("area_m2") or 0) > 0:
                ws.append([f.get("classe"), _num(f.get("area_m2")), _num(f.get("pct"))])

    # ---- Ambiental (alertas) ----
    amb = dims.get("ambiental")
    if isinstance(amb, dict):
        alertas = amb.get("alertas") or []
        ws = nova_aba("Ambiental")
        cabecalho(ws, ["Tipo", "Severidade", "Área afetada (m²)", "Detalhe"])
        for a in alertas:
            ws.append([
                a.get("tipo"), a.get("severidade"),
                _num(a.get("area_afetada_m2")), a.get("detalhe"),
            ])
        if not alertas:
            ws.append(["Sem alertas nas camadas consultadas."])

    # ---- Financeira ----
    fin = dims.get("financeira")
    if isinstance(fin, dict):
        ws = nova_aba("Financeira")
        kv(ws, "VGV bruto", _g(fin, "vgv", "bruto"))
        kv(ws, "VGV próprio", _g(fin, "vgv", "proprio"))
        kv(ws, "VGV geral", _g(fin, "vgv", "geral"))
        kv(ws, "Resultado nominal", _g(fin, "indicadores", "resultado_nominal"))
        kv(ws, "Margem s/ VGV próprio", _g(fin, "indicadores", "margem_sobre_vgv_proprio"))
        kv(ws, "Exposição máxima", _g(fin, "indicadores", "exposicao_maxima", "valor"))
        ws.append([])
        cabecalho(ws, ["Bloco de custo", "Total (R$)"])
        for b in fin.get("blocos") or []:
            ws.append([b.get("nome"), _num(b.get("total"))])
        anual = fin.get("fluxo_resumo_anual") or []
        if anual:
            ws.append([])
            cabecalho(ws, ["Ano", "Entradas", "Saídas", "Líquido", "Acumulado"])
            for r in anual:
                ws.append([
                    r.get("ano"), _num(r.get("entradas")), _num(r.get("saidas")),
                    _num(r.get("liquido")), _num(r.get("acumulado")),
                ])

    # ---- Econômica ----
    eco = dims.get("economica")
    if isinstance(eco, dict):
        ws = nova_aba("Econômica")
        kv(ws, "TMA (% a.a. real)", _g(eco, "tma", "aa_real"))
        kv(ws, "VPL", _g(eco, "vpl", "valor"))
        kv(ws, "TIR (% a.a.)", _g(eco, "tir", "aa"))
        kv(ws, "Payback simples (mês)", _g(eco, "payback", "simples_mes"))
        kv(ws, "Payback descontado (mês)", _g(eco, "payback", "descontado_mes"))
        kv(ws, "Índice de lucratividade", _g(eco, "indice_lucratividade"))

    # largura das colunas (estética leve)
    for ws in wb.worksheets:
        for col in ws.columns:
            largura = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(largura + 2, 12), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
