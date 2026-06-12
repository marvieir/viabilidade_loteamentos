#!/usr/bin/env python3
"""Pipeline (OFFLINE do runtime) — gera o arquivo embarcado da Fase 6 (Localização).

Consolida, por município + 27 UFs + Brasil:
  • população 2022 e 2010, área e densidade        — IBGE/SIDRA (Censo)
  • PIB per capita (ano mais recente do recorte)    — IBGE/SIDRA (PIB dos Municípios)
  • domicílios ocupados + moradores/domicílio       — IBGE/SIDRA (Censo 2022)
  • distribuição etária (0-14·15-29·30-59·60+)       — IBGE/SIDRA (Censo 2022)
  • déficit habitacional (quando no recorte)         — Fundação João Pinheiro (planilha local)

**Aquisição é pipeline, não agente** (ARCHITECTURE §2.5): roda na máquina do operador
(que tem egress ao IBGE), grava `backend/app/dados/localizacao_municipios.json[.gz]` com
cabeçalho de proveniência e **valida os valores-ouro de São Roque/SP antes de aceitar**.
O runtime nunca chama a rede — só lê o arquivo gerado.

Uso (no Mac do operador, a partir da raiz do repo):
    cd /caminho/para/viabilidade_loteamentos
    python3 -m pip install openpyxl   # só se for usar --fjp
    python3 scripts/gerar_localizacao_ibge.py --fjp ./dados_brutos/deficit_fjp.xlsx
    # sem --fjp: gera com deficit=null em todos (fallback de estoque assume)
    # --pib-ano: período do PIB dos Municípios (default 2023 = ano do ouro)
    # --gzip: grava .json.gz se o arquivo passar de alguns MB

Requer egress a `servicodados.ibge.gov.br`. O déficit FJP não tem API pública estável:
baixe a planilha municipal da FJP e PREPARE uma aba com as colunas `cod_ibge` (7 díg.),
`deficit` (inteiro) e `ano` — o download bruto da FJP não vem nesse layout.
"""

from __future__ import annotations

import argparse
import gzip
import json
import re
import sys
import urllib.request
from datetime import date
from pathlib import Path
from typing import Optional

API = "https://servicodados.ibge.gov.br/api/v3/agregados"
RAIZ = Path(__file__).resolve().parent.parent
DESTINO = RAIZ / "backend" / "app" / "dados" / "localizacao_municipios.json"

# Valores-ouro (São Roque/SP, 3550605, fonte IBGE) — gate de aceite da geração.
OURO_SAO_ROQUE = {
    "pop_2022": 79484,
    "pop_2010": 78821,
    "densidade": 258.98,
    "pib_per_capita": 57024.90,
    "moradores_por_domicilio": 2.79,
}


def _get(url: str) -> list | dict:
    req = urllib.request.Request(url, headers={"Accept-Encoding": "identity"})
    with urllib.request.urlopen(req, timeout=300) as resp:  # noqa: S310 (URL fixa do IBGE)
        bruto = resp.read()
    if bruto[:2] == b"\x1f\x8b":  # alguns endpoints respondem gzip mesmo pedindo identity
        bruto = gzip.decompress(bruto)
    return json.loads(bruto.decode("utf-8"))


def _variavel_por_nome(agregado: str, trecho: str, fallback: str) -> str:
    """Resolve o id da variável pelos METADADOS do agregado (não chuta id fixo).

    Caso real: no 5938, a v37 é o PIB *total* (mil R$); o per capita é a v6575 —
    e ela NÃO aparece na lista `variaveis` do metadados (variável derivada). Por isso,
    além de buscar por nome, imprimimos os candidatos e caímos num fallback conhecido.
    """
    try:
        meta = _get(f"{API}/{agregado}/metadados")
        variaveis = meta.get("variaveis", []) if isinstance(meta, dict) else []
        for v in variaveis:
            if trecho.lower() in str(v.get("nome", "")).lower():
                print(f"  variável {v['id']} = “{v['nome']}”", file=sys.stderr)
                return str(v["id"])
        candidatos = ", ".join(f"{v.get('id')}={v.get('nome')}" for v in variaveis) or "(lista vazia)"
        print(
            f"  (aviso: nenhuma variável com “{trecho}” no {agregado} → uso fallback v{fallback}. "
            f"Candidatos: {candidatos})",
            file=sys.stderr,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"  (aviso: metadados do agregado {agregado} indisponíveis: {exc}; uso v{fallback})", file=sys.stderr)
    return fallback


def _resolver_periodo(agregado: str, preferido: str) -> str:
    """Devolve `preferido` se existir no agregado; senão o período mais recente (evita 500)."""
    try:
        per = _get(f"{API}/{agregado}/periodos")
        disponiveis = sorted(str(p["id"]) for p in per) if isinstance(per, list) else []
    except Exception as exc:  # noqa: BLE001
        print(f"  (aviso: períodos do {agregado} indisponíveis: {exc}; tento {preferido})", file=sys.stderr)
        return preferido
    if not disponiveis:
        return preferido
    if preferido in disponiveis:
        return preferido
    ultimo = disponiveis[-1]
    print(
        f"  ⚠ período {preferido} indisponível no {agregado}; uso o mais recente {ultimo} "
        f"(disponíveis: …{', '.join(disponiveis[-4:])})",
        file=sys.stderr,
    )
    return ultimo


def _valores_por_localidade(agregado: str, variavel: str, periodo: str, nivel: str) -> dict[str, float]:
    """Retorna {cod_localidade: valor} para um agregado/variável/período do SIDRA v3."""
    url = f"{API}/{agregado}/periodos/{periodo}/variaveis/{variavel}?localidades={nivel}"
    dados = _get(url)
    out: dict[str, float] = {}
    for var in dados:
        for res in var.get("resultados", []):
            for serie in res.get("series", []):
                cod = serie["localidade"]["id"]
                valores = serie.get("serie", {})
                bruto = valores.get(periodo)
                if bruto not in (None, "...", "-", "X"):
                    try:
                        out[cod] = float(bruto)
                    except ValueError:
                        pass
    return out


def _coletar_niveis(agregado: str, variavel: str, periodo: str, rotulo: str) -> dict[str, float]:
    """Coleta N6+N3+N1 com resiliência POR NÍVEL: um 500 (ex.: N6 grande demais) não
    derruba o bloco nem o programa — apenas registra o aviso e segue."""
    out: dict[str, float] = {}
    for nivel in ("N6", "N3", "N1"):
        try:
            out.update(_valores_por_localidade(agregado, variavel, periodo, nivel))
        except Exception as exc:  # noqa: BLE001
            print(f"  (aviso: {rotulo} [{nivel}] falhou: {exc})", file=sys.stderr)
    return out


def coletar_ibge(pib_ano: str) -> dict[str, dict]:
    """Monta os registros por município/UF/Brasil a partir do SIDRA. Resiliente por bloco:
    cada indicador que falhar fica ausente (runtime degrada para PARCIAL), nunca aborta tudo."""
    # Níveis SIDRA: N6 = município, N3 = UF, N1 = Brasil.
    registros: dict[str, dict] = {}

    print("→ população 2022 (Censo, agregado 4714 v93)…", file=sys.stderr)
    pop22 = _coletar_niveis("4714", "93", "2022", "população 2022")
    print("→ densidade 2022 (agregado 4714 v614)…", file=sys.stderr)
    dens = _coletar_niveis("4714", "614", "2022", "densidade")
    print("→ população 2010 (Censo, agregado 1378 v93)…", file=sys.stderr)
    pop10 = _coletar_niveis("1378", "93", "2010", "população 2010")

    # PIB per capita: o agregado 5938 NÃO tem variável per capita (só PIB total = v37, em
    # mil R$ — a lista de candidatos confirmou). Então CALCULAMOS: pib_pc = v37 × 1000 ÷ pop.
    # É o caminho determinístico do projeto (cálculo no backend, não confiar em campo de API).
    pib_periodo = _resolver_periodo("5938", pib_ano)
    print(f"→ PIB total (agregado 5938 v37, {pib_periodo}) → per capita = ÷ população…", file=sys.stderr)
    pib_total = _coletar_niveis("5938", "37", pib_periodo, "PIB total")  # mil R$
    pib_ano = pib_periodo  # o ano realmente usado vai para a proveniência

    print("→ domicílios e moradores/domicílio (Censo 2022)…", file=sys.stderr)
    dom = _coletar_niveis("4712", "381", "2022", "domicílios")
    mpd = _coletar_niveis("4712", "5930", "2022", "moradores/domicílio")

    print("→ faixa etária (Censo 2022, agregado 9514)…", file=sys.stderr)
    try:
        faixas = _coletar_faixa_etaria()
    except Exception as exc:  # noqa: BLE001
        print(f"  (aviso: faixa etária falhou inteira: {exc} — fica null, runtime → PARCIAL)", file=sys.stderr)
        faixas = {}

    def _pib_pc(cod: str) -> Optional[float]:
        if cod in pib_total and cod in pop22 and pop22.get(cod):
            return round(pib_total[cod] * 1000 / pop22[cod], 2)
        return None

    todos_cods = set(pop22) | set(pop10) | set(pib_total)
    for cod in todos_cods:
        if len(cod) == 7:
            chave, nivel = cod, "municipio"
        elif len(cod) == 2:
            chave, nivel = f"UF:{_sigla_uf(cod)}", "uf"
        elif cod == "1":
            chave, nivel = "BR", "brasil"
        else:
            continue
        registros[chave] = {
            "nivel": nivel,
            "cod": cod,
            "nome": None,  # preenchido abaixo pelos metadados de localidades
            "uf": _sigla_uf(cod[:2]) if nivel == "municipio" else (_sigla_uf(cod) if nivel == "uf" else None),
            "pop_2022": int(pop22[cod]) if cod in pop22 else None,
            "pop_2010": int(pop10[cod]) if cod in pop10 else None,
            "area_km2": round(pop22[cod] / dens[cod], 3) if cod in dens and cod in pop22 and dens.get(cod) else None,
            "pib_per_capita": _pib_pc(cod),
            "pib_ano": int(pib_ano),
            "domicilios_ocupados": int(dom[cod]) if cod in dom else None,
            "moradores_por_domicilio": round(mpd[cod], 2) if cod in mpd else None,
            "deficit": None,
            "faixa_etaria": faixas.get(cod),
        }
    _nomear(registros)
    return registros


def _coletar_faixa_etaria() -> dict[str, dict]:
    """Distribuição em 4 grupos a partir das faixas quinquenais do Censo 2022 (agregado 9514).

    Retorna {cod: {"0-14":.., "15-29":.., "30-59":.., "60+":..}} com Σ=1 (normalizado pelo
    próprio total dos grupos). A classificação de idade e as categorias quinquenais são
    resolvidas pelos METADADOS (sem id chutado): pega a classificação que contém a
    categoria "0 a 4 anos" e filtra "N a N+4 anos" + "100 anos ou mais" — evita somar
    grupos sobrepostos (ex.: "0 a 14 anos") em dobro. Best-effort: falhou → {} (faixa
    etária fica indisponível; o runtime degrada honesto para PARCIAL)."""
    agregado = "9514"
    try:
        meta = _get(f"{API}/{agregado}/metadados")
    except Exception as exc:  # noqa: BLE001
        print(f"  (aviso: metadados do {agregado} indisponíveis: {exc} — faixa etária fica null)", file=sys.stderr)
        return {}

    classif = None
    for c in meta.get("classificacoes", []):
        if any(str(cat.get("nome", "")).strip() == "0 a 4 anos" for cat in c.get("categorias", [])):
            classif = c
            break
    if classif is None:
        print("  (aviso: classificação de idade não encontrada no 9514 — faixa etária fica null)", file=sys.stderr)
        return {}

    # Só as quinquenais exatas (lb..lb+4) + "100 anos ou mais"; {id_categoria: idade_inicial}.
    cats: dict[str, int] = {}
    for cat in classif["categorias"]:
        nome = str(cat.get("nome", "")).strip()
        m = re.match(r"^(\d+) a (\d+) anos$", nome)
        if m and int(m.group(2)) == int(m.group(1)) + 4:
            cats[str(cat["id"])] = int(m.group(1))
        elif re.match(r"^100 anos ou mais$", nome):
            cats[str(cat["id"])] = 100
    if len(cats) < 21:
        print(f"  (aviso: só {len(cats)} categorias quinquenais no 9514 — confira o recorte)", file=sys.stderr)

    def _grupo(lb: int) -> str:
        return "0-14" if lb < 15 else "15-29" if lb < 30 else "30-59" if lb < 60 else "60+"

    somas: dict[str, dict[str, float]] = {}
    def _acumular(dados: list) -> None:
        for var in dados:
            for res in var.get("resultados", []):
                cat_id = None  # categoria deste resultado (a chave do dict 'categoria' é o id)
                for cl in res.get("classificacoes", []):
                    if str(cl.get("id")) == str(classif["id"]):
                        cat_id = next(iter(cl.get("categoria", {})), None)
                if cat_id not in cats:
                    continue
                grupo = _grupo(cats[cat_id])
                for serie in res.get("series", []):
                    cod = serie["localidade"]["id"]
                    bruto = serie.get("serie", {}).get("2022")
                    if bruto in (None, "...", "-", "X"):
                        continue
                    try:
                        v = float(bruto)
                    except ValueError:
                        continue
                    g = somas.setdefault(cod, {"0-14": 0.0, "15-29": 0.0, "30-59": 0.0, "60+": 0.0})
                    g[grupo] += v

    ids = ",".join(cats)
    base = f"{API}/{agregado}/periodos/2022/variaveis/93"
    # N3/N1: payload pequeno (27+1 localidades) → todas as faixas numa requisição.
    for nivel in ("N3", "N1"):
        try:
            _acumular(_get(f"{base}?localidades={nivel}&classificacao={classif['id']}[{ids}]"))
        except Exception as exc:  # noqa: BLE001
            print(f"  (aviso: faixa etária [{nivel}] falhou: {exc})", file=sys.stderr)
    # N6: 5.570 municípios × 21 faixas numa tacada estoura (HTTP 500) → UMA faixa por vez.
    for cat_id in cats:
        try:
            _acumular(_get(f"{base}?localidades=N6&classificacao={classif['id']}[{cat_id}]"))
        except Exception as exc:  # noqa: BLE001
            print(f"  (aviso: faixa etária [N6 cat {cat_id}] falhou: {exc})", file=sys.stderr)

    if "1" in somas:  # sanidade visível: Σ idades Brasil deve ≈ população 2022 (203.080.756)
        total_br = int(sum(somas["1"].values()))
        print(f"  Σ idades Brasil = {total_br:,} pessoas".replace(",", "."), file=sys.stderr)

    faixas: dict[str, dict] = {}
    for cod, g in somas.items():
        total = sum(g.values())
        if total <= 0:
            continue
        faixas[cod] = {k: round(v / total, 6) for k, v in g.items()}
    return faixas


def _sigla_uf(cod2: str) -> Optional[str]:
    return _UF_POR_COD.get(cod2)


def _nomear(registros: dict[str, dict]) -> None:
    """Preenche os nomes a partir de /localidades (município e UF)."""
    try:
        muns = {str(m["id"]): m["nome"] for m in _get("https://servicodados.ibge.gov.br/api/v1/localidades/municipios")}
        ufs = {str(u["id"]): u["nome"] for u in _get("https://servicodados.ibge.gov.br/api/v1/localidades/estados")}
    except Exception as exc:  # noqa: BLE001
        print(f"  (aviso: nomes não resolvidos: {exc})", file=sys.stderr)
        return
    for chave, reg in registros.items():
        if reg["nivel"] == "municipio":
            reg["nome"] = muns.get(reg["cod"], reg["nome"])
        elif reg["nivel"] == "uf":
            reg["nome"] = ufs.get(reg["cod"], reg["nome"])
        elif reg["nivel"] == "brasil":
            reg["nome"] = "Brasil"


def aplicar_fjp(registros: dict[str, dict], caminho: Path) -> int:
    """Carrega déficit habitacional municipal da FJP (planilha) → preenche reg['deficit'].

    Espera colunas: cod_ibge (7 díg.), deficit (inteiro), ano. Município ausente fica null
    (fallback de estoque assume no runtime). NUNCA estima.
    """
    if not Path(caminho).exists():
        print(
            f"  (aviso: planilha FJP não encontrada em '{caminho}' — pulei o déficit "
            "(fica null; o fallback de estoque assume no runtime). Para incluir o déficit, "
            "baixe a planilha da FJP e aponte o caminho real com --fjp.)",
            file=sys.stderr,
        )
        return 0
    try:
        import openpyxl  # opcional; só o operador com a planilha precisa
    except ImportError:
        print(
            "  (aviso: openpyxl ausente — pulei o FJP; deficit fica null.\n"
            "   Para incluir o déficit: python3 -m pip install openpyxl  e rode de novo.)",
            file=sys.stderr,
        )
        return 0
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        print(f"  (aviso: não consegui abrir a planilha FJP ({exc}) — déficit fica null)", file=sys.stderr)
        return 0
    ws = wb.active
    cabecalho = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    try:
        i_cod = cabecalho.index("cod_ibge")
        i_def = cabecalho.index("deficit")
        i_ano = cabecalho.index("ano")
    except ValueError:
        print("  (aviso: planilha FJP sem colunas cod_ibge/deficit/ano — pulei)", file=sys.stderr)
        return 0
    n = 0
    for row in ws.iter_rows(min_row=2):
        cod = str(row[i_cod].value).split(".")[0].zfill(7) if row[i_cod].value else None
        if cod and cod in registros and row[i_def].value is not None:
            registros[cod]["deficit"] = {
                "valor": int(row[i_def].value),
                "fonte": "FJP",
                "ano": int(row[i_ano].value),
            }
            n += 1
    return n


def validar_ouro(registros: dict[str, dict]) -> None:
    """Gate: o arquivo só é aceito se São Roque bater os valores-ouro (±tolerância)."""
    reg = registros.get("3550605")
    if reg is None:
        raise SystemExit("ERRO: São Roque (3550605) ausente do recorte — não gravo.")
    densidade = round(reg["pop_2022"] / reg["area_km2"], 2) if reg.get("area_km2") else None
    # Retrato do São Roque ANTES do veredito — assim uma rodada já mostra tudo, mesmo falhando.
    print(
        "  São Roque coletado: "
        f"pop2022={reg.get('pop_2022')} pop2010={reg.get('pop_2010')} "
        f"densidade={densidade} pib_pc={reg.get('pib_per_capita')} "
        f"(ano {reg.get('pib_ano')}) morad/dom={reg.get('moradores_por_domicilio')}",
        file=sys.stderr,
    )
    checks = {
        "pop_2022": (reg.get("pop_2022"), OURO_SAO_ROQUE["pop_2022"], 0),
        "pop_2010": (reg.get("pop_2010"), OURO_SAO_ROQUE["pop_2010"], 0),
        "densidade": (densidade, OURO_SAO_ROQUE["densidade"], 0.05),
        # per capita CALCULADO (v37×1000÷pop censo) — IBGE divide pela pop estimada do ano,
        # então a tolerância é relativa (±2%), não ao centavo.
        "pib_per_capita": (reg.get("pib_per_capita"), OURO_SAO_ROQUE["pib_per_capita"], 0.02 * OURO_SAO_ROQUE["pib_per_capita"]),
        "moradores_por_domicilio": (reg.get("moradores_por_domicilio"), OURO_SAO_ROQUE["moradores_por_domicilio"], 0.01),
    }
    erros = []
    for nome, (obtido, esperado, tol) in checks.items():
        if obtido is None or abs(obtido - esperado) > tol:
            extra = ""
            if nome == "pib_per_capita" and str(reg.get("pib_ano")) != "2023":
                extra = (
                    f" — ATENÇÃO: PIB veio do ano {reg.get('pib_ano')}, não 2023 (ano do ouro). "
                    "O PIB dos Municípios pode não ter 2023 publicado; reconcilie o ano com a spec."
                )
            erros.append(f"  {nome}: obtido {obtido} ≠ ouro {esperado} (tol {tol}){extra}")
    fe = reg.get("faixa_etaria")
    if fe and abs(sum(fe.values()) - 1.0) > 0.001:
        erros.append(f"  faixa_etaria Σ = {sum(fe.values()):.4f} ≠ 1,000")
    if erros:
        raise SystemExit("ERRO: São Roque não bateu os valores-ouro — NÃO gravei:\n" + "\n".join(erros))
    if not fe:
        print(
            "⚠ AVISO: faixa etária veio VAZIA — o arquivo será aceito (degradação honesta → "
            "PARCIAL no runtime), mas confira a coleta do agregado 9514.",
            file=sys.stderr,
        )
    print("✓ valores-ouro de São Roque validados.", file=sys.stderr)


def main() -> None:
    ap = argparse.ArgumentParser(description="Gera o arquivo embarcado da Localização (Fase 6).")
    ap.add_argument("--fjp", type=Path, help="Planilha FJP do déficit habitacional municipal (.xlsx).")
    ap.add_argument("--gzip", action="store_true", help="Grava .json.gz em vez de .json.")
    ap.add_argument(
        "--pib-ano",
        default="2023",
        help="Período do PIB dos Municípios (default 2023 — é o ano do valor-ouro de São "
        "Roque; se mudar, atualize OURO_SAO_ROQUE junto, senão o gate recusa).",
    )
    ap.add_argument("--saida", type=Path, default=DESTINO)
    args = ap.parse_args()

    if args.pib_ano != "2023":
        print(
            f"⚠ AVISO: --pib-ano {args.pib_ano} ≠ 2023 (ano do ouro de São Roque). "
            "O gate vai recusar a menos que OURO_SAO_ROQUE['pib_per_capita'] seja atualizado.",
            file=sys.stderr,
        )

    registros = coletar_ibge(args.pib_ano)
    n_fjp = aplicar_fjp(registros, args.fjp) if args.fjp else 0
    validar_ouro(registros)

    dataset = {
        "_meta": {
            "data_geracao": date.today().isoformat(),
            "fontes": "IBGE — SIDRA (Censo Demográfico 2022 e 2010; PIB dos Municípios) e "
            "Fundação João Pinheiro (Déficit Habitacional Municipal)",
            "nota": f"Gerado por scripts/gerar_localizacao_ibge.py. {n_fjp} municípios com déficit FJP; "
            "os demais usam fallback de estoque de domicílios (NÃO é o déficit).",
            "niveis": ["municipio", "uf", "brasil"],
        },
        "registros": registros,
    }

    saida = args.saida
    if args.gzip:
        saida = saida.with_suffix(".json.gz")
        with gzip.open(saida, "wt", encoding="utf-8") as fh:
            json.dump(dataset, fh, ensure_ascii=False)
    else:
        saida.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ gravado {saida} ({len(registros)} registros).", file=sys.stderr)


_UF_POR_COD = {
    "11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA", "16": "AP", "17": "TO",
    "21": "MA", "22": "PI", "23": "CE", "24": "RN", "25": "PB", "26": "PE", "27": "AL",
    "28": "SE", "29": "BA", "31": "MG", "32": "ES", "33": "RJ", "35": "SP", "41": "PR",
    "42": "SC", "43": "RS", "50": "MS", "51": "MT", "52": "GO", "53": "DF",
}


if __name__ == "__main__":
    main()
